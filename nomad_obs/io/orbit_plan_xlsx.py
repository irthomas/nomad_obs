# -*- coding: utf-8 -*-
"""
Created on Mon Apr 27 12:23:44 2020

@author: iant
"""

import os
import sys
import xlsxwriter
# import xlrd
from openpyxl import load_workbook

from nomad_obs.other.generic_functions import stop
from nomad_obs.config.paths import BASE_DIRECTORY


def writeObservationPlan(worksheet, row, row_to_write):
    """function to write observation plan"""
    for column, row_item in enumerate(row_to_write):
        worksheet.write(row, column, row_item)


def writeOrbitPlanXlsx(orbit_list, mtpConstants, paths, version, place_in_base_dir=True):
    """write generic observation plan to file if the file doesn't exist """
    mtp_number = mtpConstants["mtpNumber"]

    # check if file has already been placed in correct directory. If so, don't generate again
    fileExists = os.path.isfile(os.path.join(paths["ORBIT_PLAN_PATH"], "nomad_mtp%03d_%s.xlsx" % (mtp_number, version)))

    if not fileExists:
        if place_in_base_dir:
            plan_path = os.path.join(BASE_DIRECTORY, "nomad_mtp%03d_%s.xlsx" % (mtp_number, version))
        else:
            plan_path = os.path.join(paths["ORBIT_PLAN_PATH"], "nomad_mtp%03d_%s.xlsx" % (mtp_number, version))

        with xlsxwriter.Workbook(plan_path) as workbook:
            worksheet = workbook.add_worksheet()

            if version == "plan_generic":  # find name of the orbit plan
                orbitPlanName = "genericOrbitPlanOut"
            elif version == "plan":
                orbitPlanName = "completeOrbitPlan"
            else:
                print("Error: unknown version %s. Exiting..." % version)
                stop()

            rowCounter = 0
            writeObservationPlan(worksheet, rowCounter, ["#orbitType", "irIngressHigh", "irIngressLow", "uvisIngress", "irEgressHigh",
                                 "irEgressLow", "uvisEgress", "irDayside", "uvisDayside", "irNightside", "uvisNightside", "night2dayTerminator", "comment"])

            for orbit in orbit_list:
                orbitPlan = orbit[orbitPlanName]
                row_to_write = [orbitPlan["orbitType"]]

                for genericObsType in ["irIngressHigh", "irIngressLow", "uvisIngress",
                                       "irEgressHigh", "irEgressLow", "uvisEgress",
                                       "irDayside", "uvisDayside", "irNightside", "uvisNightside"]:
                    if version == "plan_generic":  # find name of the orbit plan. Information is in a different format in the generic orbit plan
                        row_to_write.append(orbitPlan["orbitTypes"][genericObsType])
                    elif version == "plan":
                        row_to_write.append(orbitPlan[genericObsType])

                row_to_write.append(orbit["dayside"]["utcStart"])
                row_to_write.append(orbitPlan["comment"])

                rowCounter += 1
                writeObservationPlan(worksheet, rowCounter, row_to_write)
    else:
        print("File already exists, skipping generation")


def getMtpPlanXlsx(mtpConstants, paths, version):
    """read back in orbit plan after iteration by OU/BIRA"""
    mtp_number = mtpConstants["mtpNumber"]

    workbookPath = os.path.join(paths["ORBIT_PLAN_PATH"], "nomad_mtp%03d_%s.xlsx" % (mtp_number, version))

    if not os.path.exists(workbookPath):
        if version == "plan_generic":  # write message depending on which orbit plan was not found, then exit program
            print("Iterated orbit plan (%s) not found. Exiting..." % workbookPath)
        elif version == "plan":
            print("Final orbit plan (%s) not found. Exiting..." % workbookPath)
        else:
            print("Orbit plan (%s) not found. Exiting..." % workbookPath)
        sys.exit()

    # if plan found, open it and list contents
    workbook = load_workbook(workbookPath, data_only=True)
    worksheet = workbook["Sheet1"]

    mtp_plan_list = []
    mtpPlanColumns = ["orbitType", "irIngressHigh", "irIngressLow", "uvisIngress", "irEgressHigh", "irEgressLow", "uvisEgress", "irDayside",
                      "uvisDayside", "irNightside", "uvisNightside", "night2dayTerminator", "comment"]
    for row_index in range(1000):
        if row_index > 0:

            if not worksheet.cell(row_index+1, 0+1).value:
                break

            cells = {}
            for column_index in range(len(mtpPlanColumns)):
                cell = worksheet.cell(row_index+1, column_index+1).value

                if cell:
                    cells[mtpPlanColumns[column_index]] = cell
                else:
                    cells[mtpPlanColumns[column_index]] = ""

            mtp_plan_list.append(cells)

    return mtp_plan_list
