# -*- coding: utf-8 -*-
"""
Created on Fri Oct 10 15:14:08 2025

@author: iant

CALCULATE STP DATA VOLUMES FROM COP ROWS AND SUMMARY FILES

TODO:
    FIX SOLAR SCAN / MOON OBS DURATIONS
    ADD INITIALISATION TIME DATA VOLUME (PROBABLY NEGLIGIBLE)
"""


import matplotlib.pyplot as plt
from nomad_obs.config.paths import setupPaths
from nomad_obs.mtp_inputs import getMtpConstants
from nomad_obs.cop_rows.data_rates import ir_data_rate, uvis_data_rate
import os
import numpy as np

from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

mtpNumber = 100

# add the correct MTP info in obs_inputs
mtpConstants = getMtpConstants(mtpNumber)
paths = setupPaths(mtpConstants)

cop_summary_dict = {
    "dayside": {
        "xlsx_filename": "NOMAD_dayside_nadir_summary.xlsx",
        "sheet_name": "NOMAD_dayside_nadir_summary",
        "columns": range(34, 57),
        "ir_channel": "LNO",
        "ir_cop_file": "mtp%03i_ir_dayside_nadir.txt" % mtpNumber,
        "uvis_cop_file": "mtp%03i_uvis_dayside_nadir.txt" % mtpNumber,
    },

    "egress": {
        "xlsx_filename": "NOMAD_egress_solar_occulations_summary.xlsx",
        "sheet_name": "NOMAD_egress_solar_occulations_",
        "columns": range(35, 58),
        "ir_channel": "SO",
        "ir_cop_file": "mtp%03i_ir_egress_occultations.txt" % mtpNumber,
        "uvis_cop_file": "mtp%03i_uvis_egress_occultations.txt" % mtpNumber,
    },

    "grazing": {
        "xlsx_filename": "NOMAD_grazing_solar_occulations_summary.xlsx",
        "sheet_name": "NOMAD_grazing_solar_occulations",
        "columns": range(32, 55),
        "ir_channel": "SO",
        "ir_cop_file": "mtp%03i_ir_grazing_occultations.txt" % mtpNumber,
        "uvis_cop_file": "mtp%03i_uvis_grazing_occultations.txt" % mtpNumber,
    },

    "ingress": {
        "xlsx_filename": "NOMAD_ingress_and_merged_solar_occulations_summary.xlsx",
        "sheet_name": "NOMAD_ingress_and_merged_solar_",
        "columns": range(35, 58),
        "ir_channel": "SO",
        "ir_cop_file": "mtp%03i_ir_ingress_occultations.txt" % mtpNumber,
        "uvis_cop_file": "mtp%03i_uvis_ingress_occultations.txt" % mtpNumber,
    },

    "nightside": {
        "xlsx_filename": "NOMAD_nightside_nadir_summary.xlsx",
        "sheet_name": "NOMAD_nightside_nadir_summary",
        "columns": range(34, 57),
        "ir_channel": "LNO",
        "ir_cop_file": "mtp%03i_ir_nightside_nadir.txt" % mtpNumber,
        "uvis_cop_file": "mtp%03i_uvis_nightside_nadir.txt" % mtpNumber,
    },

    "calibration": {
        "xlsx_filename": "none.xlsx",
        "sheet_name": "",
        "columns": [],
        "ir_channel": "SO",
        "ir_cop_file": "mtp%03i_ir_calibrations.txt" % mtpNumber,
        "uvis_cop_file": "mtp%03i_uvis_calibration.txt" % mtpNumber,
    },

    "phobos_deimos": {
        "xlsx_filename": "none.xlsx",
        "sheet_name": "",
        "columns": [],
        "ir_channel": "SO",
        "ir_cop_file": "mtp%03i_ir_phobos_deimos.txt" % mtpNumber,
        "uvis_cop_file": "mtp%03i_uvis_phobos_deimos.txt" % mtpNumber,
    },

}


# #loop through summary files getting all generic COP row data
summary_file_data = {}
data_rates_dict = {}
dv_total = 0.0

for obs_type, dictionary_data in cop_summary_dict.items():

    # print(obs_type)

    summary_row_path = os.path.join(paths["SUMMARY_FILE_PATH"], dictionary_data["xlsx_filename"])

    # open spreadsheet
    if not os.path.exists(summary_row_path):
        if dictionary_data["xlsx_filename"] != "none.xlsx":
            print("########WARNING: %s does not exist#############" % summary_row_path)

    else:
        wb = load_workbook(summary_row_path, data_only=True)
        sheets = wb.sheetnames
        Sheet1 = wb[dictionary_data["sheet_name"]]

        summary_file_data[obs_type] = {}
        # loop through columns
        for column_number in dictionary_data["columns"]:

            # loop through rows
            for row_number in range(1000):

                # get data
                data = Sheet1.cell(row_number+1, column_number).value

                # if column header, make dict entry
                if row_number == 0:
                    summary_file_data[obs_type][data] = []
                    # save name for next row loop
                    name = data[:]

                # check if row contains data, if not skip
                elif data is not None:
                    # if data is when, add to dict
                    summary_file_data[obs_type][name].append(data)

    # calculate real durations from summary files
    # for obs_type, dictionary_data in cop_summary_dict.items():

    # name = obs_type
    ir_channel = dictionary_data["ir_channel"]

    if dictionary_data["sheet_name"] != "":

        summary_file_data[obs_type]["precooling_duration"] = [
            summary_file_data[obs_type]["TC20 %s_START_SCIENCE_1" % ir_channel][i] -
            summary_file_data[obs_type]["TC20 %s_DURATION_REFERENCE_1" % ir_channel][i] -
            summary_file_data[obs_type]["TC20 %s_START_TIME" % ir_channel][i] for i in range(len(summary_file_data[obs_type]["TC20 %s_START_SCIENCE_1" % ir_channel]))]

        # summary_file_data[obs_type]["ir_science_duration"] = [
        #     summary_file_data[obs_type]["TC20 %s_DURATION_TIME" % ir_channel][i] -
        #     summary_file_data[obs_type]["precooling_duration"][i] for i in range(len(summary_file_data[obs_type]["TC20 %s_START_SCIENCE_1" % ir_channel]))]

        summary_file_data[obs_type]["ir_science_duration"] = [
            summary_file_data[obs_type]["TC20 %s_DURATION_TIME" % ir_channel][i] for i in range(len(summary_file_data[obs_type]["TC20 %s_START_SCIENCE_1" % ir_channel]))]

        summary_file_data[obs_type]["uvis_science_duration"] = [
            summary_file_data[obs_type]["TC20 UVIS_DURATION_TIME"][i] for i in range(len(summary_file_data[obs_type]["TC20 UVIS_DURATION_TIME"]))]

    else:
        summary_file_data[obs_type] = {}

    # get real COP rows from files

    ir_cop_row_path = os.path.join(paths["COP_ROW_PATH"], dictionary_data["ir_cop_file"])
    uvis_cop_row_path = os.path.join(paths["COP_ROW_PATH"], dictionary_data["uvis_cop_file"])

    # get COP rows from file, convert to integers and save to list
    ir_cop_row_data = []
    if os.path.exists(ir_cop_row_path):
        with open(ir_cop_row_path, "r") as f:
            lines = f.readlines()

        for line in lines:
            line_split = line.split(",")
            if line_split[0] == "TC20 FIXED":
                # ir_cop_row_data.append(line_split)
                # print(line_split)
                continue
            else:
                ir_cop_row_data.append([int(x) if i in range(6) else x for i, x in enumerate(line_split)])

    summary_file_data[obs_type]["ir_cop_row_data"] = ir_cop_row_data

    # get COP rows from file, convert to integers and save to list
    uvis_cop_row_data = []
    if os.path.exists(uvis_cop_row_path):
        with open(uvis_cop_row_path, "r") as f:
            lines = f.readlines()

        for line in lines:
            line_split = line.split(",")
            if line_split[0] == "TC20 UVIS_COP_ROW" or line_split[0] == "COP row":
                # uvis_cop_row_data.append(line_split)
                # print(line_split)
                continue
            else:
                uvis_cop_row_data.append([int(x) if i in range(2) else x for i, x in enumerate(line_split)])

    summary_file_data[obs_type]["uvis_cop_row_data"] = uvis_cop_row_data

    if obs_type == "calibration":
        summary_file_data[obs_type]["precooling_duration"] = [440] * len(ir_cop_row_data)
        summary_file_data[obs_type]["ir_science_duration"] = [3600] * len(ir_cop_row_data)
        summary_file_data[obs_type]["uvis_science_duration"] = [3600] * len(uvis_cop_row_data)
    if obs_type == "phobos_deimos":
        # TODO: get this correct
        summary_file_data[obs_type]["precooling_duration"] = [440] * len(ir_cop_row_data)
        summary_file_data[obs_type]["ir_science_duration"] = [3600] * len(ir_cop_row_data)
        summary_file_data[obs_type]["uvis_science_duration"] = [3600] * len(uvis_cop_row_data)

    # get IR and UVIS data rate
    data_rates_dict[obs_type] = {}

    ir_cop_row_data = summary_file_data[obs_type]["ir_cop_row_data"]
    ir_science_duration = summary_file_data[obs_type]["ir_science_duration"]
    ir_data_rates = []
    ir_cumul_data_rates = []

    uvis_cop_row_data = summary_file_data[obs_type]["uvis_cop_row_data"]
    uvis_science_duration = summary_file_data[obs_type]["uvis_science_duration"]
    uvis_data_rates = []
    uvis_cumul_data_rates = []
    # TODO: include initialisation time data volumes / SINBAD data volume

    times = []
    dts = []
    if uvis_cop_row_data and (len(ir_cop_row_data) != len(uvis_cop_row_data)):
        print("Error: wrong length")

    for i in range(len(ir_cop_row_data)):
        tc20_fixed = ir_cop_row_data[i][0]

        if uvis_cop_row_data:
            tc20_uvis = uvis_cop_row_data[i][0]
        if tc20_fixed > -1:
            duration = ir_science_duration[i]
            if duration == 0:
                print("Error")
            else:
                ir_rate = ir_data_rate(obs_type, duration)
                ir_data_rates.append(ir_rate)
        else:
            ir_data_rates.append(0.0)

        # record obs time
        times.append(ir_cop_row_data[i][7])
        # 2025 SEP 27 16:16:53 or 2025-11-08T04:54:51
        try:
            dt = datetime.strptime(ir_cop_row_data[i][7], "%Y %b %d %H:%M:%S")
        except ValueError:
            dt = datetime.strptime(ir_cop_row_data[i][7], "%Y-%m-%dT%H:%M:%S")
        dts.append(dt)

        ir_cumul_data_rates.append(sum(ir_data_rates))

        if uvis_cop_row_data:
            if tc20_uvis > -1:
                duration = uvis_science_duration[i]
                if duration == 0:
                    print("Error")
                else:
                    uvis_rate = uvis_data_rate(tc20_uvis, duration)
                    uvis_data_rates.append(uvis_rate)
            else:
                uvis_data_rates.append(0.0)

            uvis_cumul_data_rates.append(sum(uvis_data_rates))

    data_rates_dict[obs_type]["utc_times"] = times
    data_rates_dict[obs_type]["dts"] = dts
    data_rates_dict[obs_type]["ir_data_rates"] = np.asarray(ir_data_rates)
    data_rates_dict[obs_type]["ir_cumul_data_rates"] = np.asarray(ir_cumul_data_rates)
    if uvis_cop_row_data:
        data_rates_dict[obs_type]["uvis_data_rates"] = np.asarray(uvis_data_rates)
        data_rates_dict[obs_type]["uvis_cumul_data_rates"] = np.asarray(uvis_cumul_data_rates)

    if len(data_rates_dict[obs_type]["ir_cumul_data_rates"]) > 0:
        print(obs_type, ":", data_rates_dict[obs_type]["ir_cumul_data_rates"][-1] / 1.0e6)
        dv_total += data_rates_dict[obs_type]["ir_cumul_data_rates"][-1] / 1.0e6

print("Total MTP data volume:", dv_total)
# stop()


# get OCM events from file
summary_row_path = os.path.join(paths["SUMMARY_FILE_PATH"], "OCM_events.txt")
with open(summary_row_path, "r") as f:
    lines = f.readlines()
ocm_times = [line.strip() for line in lines[2::2]]
ocm_dts = [datetime.strptime(t, "%Y-%m-%dT%H:%M:%S") for t in ocm_times]


# fig, ax = plt.subplots()
# ax.bar(data_rates_dict["dayside"]["utc_times"], data_rates_dict["dayside"]["uvis_data_rates"])

fig, ax = plt.subplots(figsize=(12, 6), constrained_layout=True)
for i, obs_type in enumerate(data_rates_dict.keys()):
    ax.plot(data_rates_dict[obs_type]["dts"], data_rates_dict[obs_type]["ir_cumul_data_rates"] /
            1.0e6, marker=".", color="C%i" % i, label="IR %s" % obs_type, linestyle="--")
    if uvis_cop_row_data:
        ax.plot(data_rates_dict[obs_type]["dts"], data_rates_dict[obs_type]["uvis_cumul_data_rates"] /
                1.0e6, marker=".", color="C%i" % i, label="UVIS %s" % obs_type, linestyle="-")
plt.legend()
plt.ylabel("Cumulative data volume Gbits")
plt.xlabel("Time")
plt.grid()
plt.title("NOMAD data generation MTP%03i" % mtpNumber)


for ocm_dt in ocm_dts:
    plt.axvline(ocm_dt, linestyle="--", color="k")

plt.savefig(os.path.join(paths["SQL_INI_PATH"], "mtp%03i_%s.png" % (mtpNumber, datetime.strftime(datetime.now(), "%Y%m%d_%H%M%S"))))
