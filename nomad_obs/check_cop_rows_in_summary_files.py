# -*- coding: utf-8 -*-
"""
Created on Tue Dec  1 12:26:54 2020

@author: iant

COPY COP ROWS TO SUMMARY FILES AND COMPARE NUMBER OF ENTRIES AND DATETIMES

"""

import os


from openpyxl import load_workbook
from openpyxl.styles import PatternFill

mtpNumber = 62

#add the correct MTP info in obs_inputs
from nomad_obs.mtp_inputs import getMtpConstants
from nomad_obs.config.paths import setupPaths



mtpConstants = getMtpConstants(mtpNumber)
paths = setupPaths(mtpConstants)

cop_summary_dict = {
    "mtp%03i_ir_dayside_nadir.txt" %mtpNumber:{
        "xlsx_filename":"NOMAD_dayside_nadir_summary.xlsx",
        "sheet_name":"NOMAD_dayside_nadir_summary",
        "empty_column":58,
        "column_to_compare":11},

    "mtp%03i_ir_egress_occultations.txt" %mtpNumber:{
        "xlsx_filename":"NOMAD_egress_solar_occulations_summary.xlsx",
        "sheet_name":"NOMAD_egress_solar_occulations_",
        "empty_column":59,
        "column_to_compare":18},

    "mtp%03i_ir_grazing_occultations.txt" %mtpNumber:{
        "xlsx_filename":"NOMAD_grazing_solar_occulations_summary.xlsx",
        "sheet_name":"NOMAD_grazing_solar_occulations",
        "empty_column":56,
        "column_to_compare":19},

    "mtp%03i_ir_ingress_occultations.txt" %mtpNumber:{
        "xlsx_filename":"NOMAD_ingress_and_merged_solar_occulations_summary.xlsx",
        "sheet_name":"NOMAD_ingress_and_merged_solar_",
        "empty_column":59,
        "column_to_compare":18},

    "mtp%03i_ir_nightside_nadir.txt" %mtpNumber:{
        "xlsx_filename":"NOMAD_nightside_nadir_summary.xlsx",
        "sheet_name":"NOMAD_nightside_nadir_summary",
        "empty_column":58,
        "column_to_compare":11},
}


#loop through summary files
for cop_row_name, dictionary_data in cop_summary_dict.items():
    
    print("Adding %s COP rows to xlsx" %dictionary_data["sheet_name"])

    cop_row_path = os.path.join(paths["COP_ROW_PATH"], cop_row_name)
    summary_row_path = os.path.join(paths["SUMMARY_FILE_PATH"], dictionary_data["xlsx_filename"])

    #get COP rows from file, convert to integers and save to list
    with open(cop_row_path, "r") as f:
        lines = f.readlines()
        
    cop_row_data = []
    for line in lines:
        line_split = line.split(",")
        if line_split[0] == "TC20 FIXED":
            cop_row_data.append(line_split)
        else:
            cop_row_data.append([int(x) if i in range(6) else x for i,x in enumerate(line_split)])
        
    
    #open spreadsheet
    if not os.path.exists(summary_row_path):
        print("########WARNING: %s does not exist#############" %summary_row_path)

    else:
        wb = load_workbook(summary_row_path, data_only=True)
        sheets = wb.sheetnames
        Sheet1 = wb[dictionary_data["sheet_name"]]
        
        #choose some rows to compare
        if len(cop_row_data) > 4:
            compare_indices = list(range(1, len(cop_row_data), int(len(cop_row_data)/5))) + [len(cop_row_data)-1]
        elif len(cop_row_data) == 1: #if no data
            compare_indices = []
        else: #print all lines
            compare_indices = list(range(1, len(cop_row_data)))
            
    
        #count number of entries in summary file
        contains_data = []
        for row_number in range(1000):
            #check if summary file row contains data
            contains_data.append(Sheet1.cell(row_number+1, dictionary_data["empty_column"]-1).value != None)
        n_rows_summary_file = sum(contains_data)
        
        #compare
        if n_rows_summary_file == len(cop_row_data):
            print("Number of rows match")
        else:
            print("Error: number of rows do not match")
    
        
        #add COP row data to summary file, one row at a time
        for row_number in range(len(cop_row_data)):
        
            #get row colour
            #note: if the colour is set with a theme this won't work
            rgb_or_theme = ""
            cell_colour_info = Sheet1.cell(row_number+1, dictionary_data["empty_column"]-1).fill.start_color
            if cell_colour_info.type == "rgb":
                rgb_or_theme = "rgb"
                color_in_hex = Sheet1.cell(row_number+1, dictionary_data["empty_column"]-1).fill.start_color.index
            else:
                rgb_or_theme = "theme"
                color_theme = Sheet1.cell(row_number+1, dictionary_data["empty_column"]-1).fill.start_color.theme
            
            #loop through COP row values
            for column_number in range(len(cop_row_data[0])):
                column_index = column_number + dictionary_data["empty_column"]
                
                #copy COP row values to cells
                Sheet1.cell(row_number+1, column_index+1).value = cop_row_data[row_number][column_number]
        
                #if row has a colour, copy to each cell
                if rgb_or_theme == "rgb":
                    if color_in_hex != '00000000':
                        fill_pattern = PatternFill(start_color=color_in_hex, end_color=color_in_hex, fill_type='solid')
                        Sheet1.cell(row_number+1, column_index+1).fill = fill_pattern
                if rgb_or_theme == "theme":
                    Sheet1.cell(row_number+1, column_index+1).fill.start_color.theme = color_theme
                    
            
            #print comparison rows TC execution times
            if row_number in compare_indices:
                print(cop_row_data[row_number][7], "---", Sheet1.cell(row_number+1, dictionary_data["column_to_compare"]).value)
    
            if rgb_or_theme == "rgb":
                if color_in_hex != '00000000':
                    if cop_row_data[row_number][0] > -1:
                        print("Error: row %i contains observation data" %(row_number+1))
                        print(cop_row_data[row_number])
            if rgb_or_theme == "theme":
                if cop_row_data[row_number][0] > -1:
                    print("Error: row %i contains observation data" %(row_number+1))
                    print(cop_row_data[row_number])
        
        #save and close file
        wb.save(summary_row_path) 







