# -*- coding: utf-8 -*-
"""
Created on Wed May 11 10:17:41 2022

@author: iant

PARSE ORBIT PLAN AND REMOVE SELECTED NADIRS

TODO:
    SET IR REGIONS OF INTEREST TO GENERIC OBSERVATION irDayside
    MAKE INTO FUNCTION TO BE RUN BY MAIN SCRIPT (MUST NOT CONTINUE EXECUTION AFTER)
    

"""
import os
import shutil
import numpy as np
from openpyxl import load_workbook

from nomad_obs.mtp_inputs import getMtpConstants
from nomad_obs.config.paths import setupPaths, BASE_DIRECTORY

from nomad_obs.regions_of_interest import nadirRegionsOfInterest

from nomad_obs.io.orbit_plan_xlsx import getMtpPlanXlsx


mtpNumber = 102
mtpConstants = getMtpConstants(mtpNumber)
paths = setupPaths(mtpConstants)

# percent of LNO limb/nadir in this MTP
GOAL_PERCENT = 10.0


# superceded by required_dayside_orbits from mtp_inputs.py for joint obs or special regions
REGIONS_TO_ALWAYS_RUN = [
    # "&daysideMatch:CERAUNIUS THOLUS",
    # "&daysideMatch:ULYSSES THOLUS",
    # "&daysideMatch:ELYSIUM MONS",
    # "&daysideMatch:ARSIA MONS",
    # "&daysideMatch:PAVONIS MONS",
    # "&daysideMatch:ASCRAEUS MONS",
    # "&daysideMatch:OLYMPUS MONS",

]

# nudge nadirs towards orbits with a single or two solar occultations
# ingress, egress, both, merged_grazing
# TODO: add none
PRIORITY_TYPES = ["ingress", "egress"]
REDUCTION_TYPES = ["both", "merged_grazing"]

MEASURE_IR_DAY_LIMBS = False
MEASURE_IR_NIGHT_NADIR_OR_LIMBS = False

VAL_OFF = -20
VAL_LIMB = 9
VAL_REQ = 10
VAL_MRO = 2
VAL_PRIORITY = 0.5
VAL_REDUCTION = -5.0


forbidden_dayside_orbits = mtpConstants["forbidden_dayside_orbits"]
required_dayside_orbits = mtpConstants["required_dayside_orbits"]


def get_adjacent_indices(indices, n_orbits, n_adjacents):
    out = []
    for index in indices:
        adj_ixs = np.arange(index - n_adjacents, index + n_adjacents + 1)
        ix = np.where((adj_ixs >= 0) & (adj_ixs < n_orbits))[0]
        out.extend(list(adj_ixs[ix]))

    return sorted(list(set(out)))


# copy generic plan to orbit_plans
src = os.path.join(BASE_DIRECTORY, "nomad_mtp%03d_%s.xlsx" % (mtpNumber, "plan_generic"))
dest = os.path.join(paths["ORBIT_PLAN_PATH"], "nomad_mtp%03d_%s.xlsx" % (mtpNumber, "plan_generic"))
shutil.copyfile(src, dest)


# get data from spreadsheet
mtpPlan = getMtpPlanXlsx(mtpConstants, paths, "plan_generic")

irDaysides = [d["irDayside"] for d in mtpPlan]
irNightsides = [d["irNightside"] for d in mtpPlan]
ingresses = [d["irIngressHigh"] for d in mtpPlan]
egresses = [d["irEgressHigh"] for d in mtpPlan]
ots = [d["orbitType"] for d in mtpPlan]
comments = [d["comment"] for d in mtpPlan]


n_orbits = len(comments)
goal_orbits = int(np.floor(n_orbits / GOAL_PERCENT))

"""orbit mask values
1st column - definite on/off
-10 = Off
9 = limb (if measured)
10 = Required dayside + regions to always run

2nd column - optional daysides
+2 = MROs
+0<1 = regions of interest
+0.5 for short/long preference

3rd column - negatives
-0.5 for reduced preference daysides
e.g. to stop consecutive measurements

4th column - sum of columns
updated each nadir placement

last column - decision on or off
0 = Off
1 = On

"""
orbit_mask = np.zeros((n_orbits, 5))

"""add required daysides and limbs, 1st column"""
# get indices of required_dayside_orbits from mtp_inputs.py
indices_required = [i-1 for i in required_dayside_orbits]
orbit_mask[indices_required, 0] = VAL_REQ

indices_day_limbs = [i for i, s in enumerate(irDaysides) if "irLimb" in s]
if MEASURE_IR_DAY_LIMBS:
    orbit_mask[indices_day_limbs, 0] = VAL_LIMB


# add nadir regions - superceded by required_dayside_orbits from mtp_inputs.py
region_comments = ["&daysideMatch:%s" % s[0] for s in nadirRegionsOfInterest]
region_ratios = [s[2] for s in nadirRegionsOfInterest]

for i, comment in enumerate(comments):
    # loop through regions of interest, checking comment for matching names
    for region_comment, region_ratio in zip(region_comments, region_ratios):
        # if found
        if region_comment in comment:
            # override to always run certain regions
            if region_comment in REGIONS_TO_ALWAYS_RUN:
                orbit_mask[i, 0] = VAL_REQ
            else:
                # add regions of interest scaler
                orbit_mask[i, 1] = region_ratio


"""add forbidden 1st column"""
# find indices of orbits where daysides are not allowed
# during OCMs
# during ACS calibrations
# during NOMAD calibrations

# before and after nightside limbs
# before and after CaSSIS calibrations
# before and after Phobos/Deimos tracking

indices_night_limbs = [i for i, s in enumerate(irNightsides) if "irNightLimb" in s]
indices_nomad_cals = [i for i, s in enumerate(comments) if "&nomadSolarCalibration" in s]
indices_acs_cals = [i for i, s in enumerate(comments) if "&acsSolarCalibration" in s]
indices_cassis_cals = [i for i, s in enumerate(comments) if "&cassisSolarCalibration" in s]
indices_phobos = [i for i, s in enumerate(comments) if "&nomadPhobos" in s]
indices_deimos = [i for i, s in enumerate(comments) if "&nomadDeimos" in s]

# get list of orbits to avoid
# for OCMs, can just avoid the orbit of the OCM
# for others, take +-1 orbit margin. Phobos, Deimos, cals must be set to -1 (programmed elsewhere)
indices_dict = {
    "ocm": [i for i, s in enumerate(comments) if "&possibleOCM" in s],
    "acs_cals": get_adjacent_indices(indices_acs_cals, n_orbits, 1),
    "night_limbs": get_adjacent_indices(indices_night_limbs, n_orbits, 1),
    "cassis_cals": get_adjacent_indices(indices_cassis_cals, n_orbits, 1),
    "phobos": get_adjacent_indices(indices_phobos, n_orbits, 1),
    "deimos": get_adjacent_indices(indices_deimos, n_orbits, 1),
    "nomad_cals": get_adjacent_indices(indices_nomad_cals, n_orbits, 1),
}


if not MEASURE_IR_DAY_LIMBS:
    # to set them all off if not measured
    indices_dict["day_limbs"] = get_adjacent_indices(indices_day_limbs, n_orbits, 1)


# collect all definite off orbits in one list of orbit numbers
indices_off = []
for key, key_indices in indices_dict.items():
    for index in key_indices:
        if index not in indices_off:
            indices_off.append(index)
indices_off = sorted(indices_off)

# set orbit masks to -1 (not allowed)
orbit_mask[indices_off, 0] = VAL_OFF

indices_forbidden = [i-1 for i in forbidden_dayside_orbits]
orbit_mask[indices_forbidden, 0] = VAL_OFF


"""add MRO overlaps 2nd column"""
# occ_indices = [i for i, orbit in enumerate(mtpPlan) if orbit["orbitType"] in [1, 5]]
# indices_mro = [i for i, s in enumerate(comments) if "&mroOverlap" in s and i not in occ_indices]
indices_mro = [i for i, s in enumerate(comments) if "&mroOverlap" in s]
orbit_mask[indices_mro, 1] += VAL_MRO


"""add short/long dayside priority"""
# find indices of orbits with different types of occultations
indices_ingresses_egresses = [i for i, (si, se) in enumerate(zip(ingresses, egresses)) if si != "" and se != ""]
indices_ingresses = [i for i, (s, ot) in enumerate(zip(ingresses, ots)) if s != "" and ot == 1 and i not in indices_ingresses_egresses]
indices_egresses = [i for i, (s, ot) in enumerate(zip(egresses, ots)) if s != "" and ot == 1 and i not in indices_ingresses_egresses]
indices_merged_grazing = [i for i, (s, ot) in enumerate(zip(ingresses, ots)) if s != "" and ot == 5]


# if PRIORITY TYPE given
if "ingress" in PRIORITY_TYPES:
    orbit_mask[indices_ingresses, 1] += VAL_PRIORITY
if "egress" in PRIORITY_TYPES:
    orbit_mask[indices_egresses, 1] += VAL_PRIORITY
if "both" in PRIORITY_TYPES:
    orbit_mask[indices_ingresses_egresses, 1] += VAL_PRIORITY
if "merged_grazing" in PRIORITY_TYPES:
    orbit_mask[indices_merged_grazing, 1] += VAL_PRIORITY

# if REDUCTION TYPE given
if "ingress" in REDUCTION_TYPES:
    orbit_mask[indices_ingresses, 2] += VAL_REDUCTION
if "egress" in REDUCTION_TYPES:
    orbit_mask[indices_egresses, 2] += VAL_REDUCTION
if "both" in REDUCTION_TYPES:
    orbit_mask[indices_ingresses_egresses, 2] += VAL_REDUCTION
if "merged_grazing" in REDUCTION_TYPES:
    orbit_mask[indices_merged_grazing, 2] += VAL_REDUCTION


# now loop through, summing totals and assigning nadirs to best orbits
for i in range(goal_orbits):

    # sum up
    orbit_mask[:, 3] = np.sum(orbit_mask[:, 0:3], axis=1)
    best_orbit_ix = np.argsort(orbit_mask[:, 3])[-1]
    orbit_mask[best_orbit_ix, -1] = 1

    # set negatives around
    ixs = get_adjacent_indices([best_orbit_ix], n_orbits, 3)
    orbit_mask[ixs, 2] -= 10


# now overwrite irDaysides with new list
nadirs = []
count = 0.0
for i in range(n_orbits):
    if orbit_mask[i, 0] == VAL_LIMB:
        nadirs.append("irLimb")

    if orbit_mask[i, -1] == 1:

        if irDaysides[i] == "":
            nadirs.append("irDayside")
            count += 1.0
        else:
            nadirs.append(irDaysides[i])
            count += 1.0
    else:
        nadirs.append("")


print("Ratio of dayside limbs/nadirs in MTP: %0.1f%%" % (100.*count/n_orbits))


# update irDaysides in spreadsheet
wb = load_workbook(dest, data_only=True)
sheets = wb.sheetnames
Sheet1 = wb["Sheet1"]

row_number = 0

for i, nadir in enumerate(nadirs):
    # copy to cell
    Sheet1.cell(row_number+2, 7+1).value = nadir  # add 2, one for header, one for 1-indexing

    # change orbit type 3 to 14 if no nadir in 1st column
    if nadir == "":
        if Sheet1.cell(row_number+2, 0+1).value == 3:
            Sheet1.cell(row_number+2, 0+1).value = 14

    # change orbit type 14 to 3 if nadir in 1st column
    if nadir != "":
        if Sheet1.cell(row_number+2, 0+1).value == 14:
            Sheet1.cell(row_number+2, 0+1).value = 3

    # if dayside nadir not allowed
    if i in indices_forbidden:

        # if orbit type 3, then set to 14
        if Sheet1.cell(row_number+2, 0+1).value in [3]:
            Sheet1.cell(row_number+2, 0+1).value = 14
        # if dayside, remove
        if Sheet1.cell(row_number+2, 8+1).value == "uvisDayside":
            Sheet1.cell(row_number+2, 8+1).value = ""

    # remove everything in the irNightside column
    if not MEASURE_IR_NIGHT_NADIR_OR_LIMBS and row_number > 0:
        Sheet1.cell(row_number+2, 9+1).value = ""

    row_number += 1


# save and close file
wb.save(dest)
print("Generic orbit plan saved to %s" % dest)
